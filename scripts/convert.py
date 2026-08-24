import json
import re
from urllib.parse import unquote
import requests
from openpyxl import load_workbook

INPUT_FILE = "Debi-Dorshon.xlsx"
OUTPUT_FILE = "debi_dorshon.json"


def extract_coordinates(url, session):
    """
    Extract latitude/longitude from Google Maps short/redirect URLs using
    URL patterns or HTML staticmap metadata fallbacks.
    """
    try:
        response = session.get(
            url,
            allow_redirects=True,
            timeout=10
        )

        final_url = unquote(response.url)

        # 1. Pattern !3d<lat>!4d<lng>
        match = re.search(r'!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)', final_url)
        if match:
            return float(match.group(1)), float(match.group(2))

        # 2. Pattern @<lat>,<lng>
        match = re.search(r'@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)', final_url)
        if match:
            return float(match.group(1)), float(match.group(2))

        # 3. staticmap center=<lat>%2C<lng> in HTML content
        match = re.search(r'center=(-?\d+(?:\.\d+)?)%2C(-?\d+(?:\.\d+)?)', response.text)
        if match:
            return float(match.group(1)), float(match.group(2))

        # 4. staticmap center=<lat>,<lng> unescaped
        match = re.search(r'center=(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)', response.text)
        if match:
            return float(match.group(1)), float(match.group(2))

        # 5. [null,null,lat,lng] JS array structure
        match = re.search(r'\[null,null,(-?\d+\.\d+),(-?\d+\.\d+)\]', response.text)
        if match:
            return float(match.group(1)), float(match.group(2))

        print("  ❌ Coordinates not found")

    except Exception as e:
        print(f"  ❌ Error fetching coordinates: {e}")

    return None, None


def parse_metro(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ["nil", "none"]:
        return None

    match = re.match(r'^(.*?)(?:\s*\((.*?)\))?$', s)
    if match:
        name = match.group(1).strip()
        line = match.group(2).strip() if match.group(2) else None
        res = {"name": name}
        if line:
            res["line"] = line
        return res

    return {"name": s}


def parse_station(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ["nil", "none"]:
        return None
    return {"name": s}


def main():
    print("Loading workbook...")
    workbook = load_workbook(INPUT_FILE, data_only=True)

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9"
    })

    all_pandals = []
    total_links = 0
    successful = 0
    failed = 0

    for sheet_name in workbook.sheetnames:
        print(f"\n==============================")
        print(f"Processing sheet: {sheet_name}")
        print(f"==============================")

        sheet = workbook[sheet_name]

        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                val_str = str(sheet.cell(r, c).value or "").strip()

                if val_str in ["Sl.no.", "Sl. No.", "Sl.No."]:
                    c_sl = c
                    c_pandal = c + 1
                    c_loc = c + 2
                    c_lat = c + 3
                    c_lng = c + 4
                    c_metro = c + 5
                    c_station = c + 6
                    c_ferry = c + 7

                    # Extract cluster name
                    cluster_name = None
                    search_r = r - 1
                    while search_r >= 1:
                        top_val = sheet.cell(search_r, c).value
                        if top_val:
                            cluster_name = str(top_val).strip()
                            break
                        search_r -= 1

                    if not cluster_name:
                        cluster_name = sheet_name

                    # Parse data rows
                    data_r = r + 1
                    while data_r <= sheet.max_row:
                        sl_val = sheet.cell(data_r, c_sl).value
                        pandal_name_val = sheet.cell(data_r, c_pandal).value

                        # Stop if sl_val and pandal_name are both empty or if sl_val is header
                        if (sl_val is None and pandal_name_val is None) or str(sl_val or "").strip() in ["Sl.no.", "Sl. No."]:
                            break

                        pandal_name = str(pandal_name_val).strip() if pandal_name_val is not None else ""
                        if not pandal_name:
                            data_r += 1
                            continue

                        loc_cell = sheet.cell(data_r, c_loc)

                        latitude = None
                        longitude = None

                        sheet_lat = sheet.cell(data_r, c_lat).value if c_lat <= sheet.max_column else None
                        sheet_lng = sheet.cell(data_r, c_lng).value if c_lng <= sheet.max_column else None

                        if isinstance(sheet_lat, (int, float)) and isinstance(sheet_lng, (int, float)):
                            latitude = float(sheet_lat)
                            longitude = float(sheet_lng)
                        elif loc_cell.hyperlink and loc_cell.hyperlink.target:
                            url = loc_cell.hyperlink.target
                            if "google" in url or "goo.gl" in url:
                                total_links += 1
                                print(f"Found link for [{cluster_name}] '{pandal_name}': {url}")
                                latitude, longitude = extract_coordinates(url, session)
                                if latitude is not None:
                                    successful += 1
                                else:
                                    failed += 1

                        metro_val = sheet.cell(data_r, c_metro).value if c_metro <= sheet.max_column else None
                        station_val = sheet.cell(data_r, c_station).value if c_station <= sheet.max_column else None
                        ferry_val = sheet.cell(data_r, c_ferry).value if c_ferry <= sheet.max_column else None

                        pandal_obj = {
                            "name": pandal_name,
                            "region": sheet_name,
                            "cluster": cluster_name,
                            "location": {
                                "latitude": latitude,
                                "longitude": longitude
                            }
                        }

                        metro_obj = parse_metro(metro_val)
                        if metro_obj:
                            pandal_obj["nearest_metro"] = metro_obj

                        station_obj = parse_station(station_val)
                        if station_obj:
                            pandal_obj["nearest_station"] = station_obj

                        ferry_obj = parse_station(ferry_val)
                        if ferry_obj:
                            pandal_obj["nearest_ferry"] = ferry_obj

                        all_pandals.append(pandal_obj)
                        data_r += 1

    # Save JSON
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(all_pandals, f, indent=2, ensure_ascii=False)

    print("\n" + "=" * 40)
    print("           FINAL SUMMARY")
    print("=" * 40)
    print(f"Total Pandals Processed: {len(all_pandals)}")
    print(f"Maps links found       : {total_links}")
    print(f"Coordinates extracted  : {successful}")
    print(f"Coordinates failed     : {failed}")
    print("=" * 40)
    print(f"\nJSON created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()