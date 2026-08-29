import json
import os
import re
from typing import Any
from urllib.parse import unquote
import requests
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(BASE_DIR, ".."))
INPUT_FILE = os.path.join(PROJECT_ROOT, "data", "raw", "Debi-Dorshon.xlsx")
OUTPUT_FILE = os.path.join(PROJECT_ROOT, "data", "processed", "debi_dorshon.json")
CACHE_FILE = os.path.join(PROJECT_ROOT, "data", "processed", "url_cache.json")


def load_cache():
    """Load URL-to-coordinates cache from file if it exists."""
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"⚠️ Warning: Could not read cache file ({e}). Starting with fresh cache.")
    return {}


def save_cache(cache):
    """Save URL-to-coordinates cache to file."""
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ Warning: Could not save cache file ({e}).")


def extract_coordinates(url, session, cache):
    """
    Extract latitude/longitude from Google Maps short/redirect URLs.
    Checks memory/file cache first. Only makes HTTP requests if the URL is new.
    """
    # 1. Check cache first
    if url in cache and cache[url] is not None:
        cached_loc = cache[url]
        if isinstance(cached_loc, dict) and "latitude" in cached_loc and "longitude" in cached_loc:
            return cached_loc["latitude"], cached_loc["longitude"], True

    # 2. Fetch via network if not in cache
    try:
        response = session.get(
            url,
            allow_redirects=True,
            timeout=10
        )

        final_url = unquote(response.url)

        # Pattern 1: !3d<lat>!4d<lng>
        match = re.search(r'!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)', final_url)
        if match:
            lat, lng = float(match.group(1)), float(match.group(2))
            cache[url] = {"latitude": lat, "longitude": lng}
            return lat, lng, False

        # Pattern 2: @<lat>,<lng>
        match = re.search(r'@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)', final_url)
        if match:
            lat, lng = float(match.group(1)), float(match.group(2))
            cache[url] = {"latitude": lat, "longitude": lng}
            return lat, lng, False

        # Pattern 3: staticmap center=<lat>%2C<lng> in HTML content
        match = re.search(r'center=(-?\d+(?:\.\d+)?)%2C(-?\d+(?:\.\d+)?)', response.text)
        if match:
            lat, lng = float(match.group(1)), float(match.group(2))
            cache[url] = {"latitude": lat, "longitude": lng}
            return lat, lng, False

        # Pattern 4: staticmap center=<lat>,<lng> unescaped
        match = re.search(r'center=(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)', response.text)
        if match:
            lat, lng = float(match.group(1)), float(match.group(2))
            cache[url] = {"latitude": lat, "longitude": lng}
            return lat, lng, False

        # Pattern 5: [null,null,lat,lng] JS array structure
        match = re.search(r'\[null,null,(-?\d+\.\d+),(-?\d+\.\d+)\]', response.text)
        if match:
            lat, lng = float(match.group(1)), float(match.group(2))
            cache[url] = {"latitude": lat, "longitude": lng}
            return lat, lng, False

        print("  ❌ Coordinates not found")

    except Exception as e:
        print(f"  ❌ Error fetching coordinates: {e}")

    return None, None, False


def parse_metro(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ["nil", "none"]:
        return None

    match = re.match(r'^(.*?)(?:\s*\((.*?)\))?$', s)
    if match:
        name = match.group(1).strip()
        line = match.group(2).strip() if match.group(2) else None
        res = {"name": name}
        if line:
            res["line"] = line
        return res

    return {"name": s}


def parse_station(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ["nil", "none"]:
        return None
    return {"name": s}


def main():
    print("Loading workbook...")
    workbook = load_workbook(INPUT_FILE, data_only=True)

    cache = load_cache()
    print(f"Loaded {len(cache)} cached coordinates from {CACHE_FILE}")

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9"
    })

    all_pandals = []
    total_links = 0
    cached_count = 0
    fetched_count = 0
    failed_count = 0

    for sheet_name in workbook.sheetnames:
        print(f"\n==============================")
        print(f"Processing sheet: {sheet_name}")
        print(f"==============================")

        sheet = workbook[sheet_name]

        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                val_str = str(sheet.cell(r, c).value or "").strip()

                if val_str in ["Sl.no.", "Sl. No.", "Sl.No."]:
                    c_sl = c
                    c_pandal = c + 1
                    c_loc = c + 2
                    c_lat = c + 3
                    c_lng = c + 4
                    c_metro = c + 5
                    c_station = c + 6
                    c_ferry = c + 7

                    # Extract cluster name
                    cluster_name = None
                    search_r = r - 1
                    while search_r >= 1:
                        top_val = sheet.cell(search_r, c).value
                        if top_val:
                            cluster_name = str(top_val).strip()
                            break
                        search_r -= 1

                    if not cluster_name:
                        cluster_name = sheet_name

                    # Parse data rows
                    data_r = r + 1
                    while data_r <= sheet.max_row:
                        sl_val = sheet.cell(data_r, c_sl).value
                        pandal_name_val = sheet.cell(data_r, c_pandal).value

                        # Stop if sl_val and pandal_name are both empty or if sl_val is header
                        if (sl_val is None and pandal_name_val is None) or str(sl_val or "").strip() in ["Sl.no.", "Sl. No."]:
                            break

                        pandal_name = str(pandal_name_val).strip() if pandal_name_val is not None else ""
                        if not pandal_name:
                            data_r += 1
                            continue

                        loc_cell = sheet.cell(data_r, c_loc)

                        latitude = None
                        longitude = None

                        sheet_lat = sheet.cell(data_r, c_lat).value if c_lat <= sheet.max_column else None
                        sheet_lng = sheet.cell(data_r, c_lng).value if c_lng <= sheet.max_column else None

                        if isinstance(sheet_lat, (int, float)) and isinstance(sheet_lng, (int, float)):
                            latitude = float(sheet_lat)
                            longitude = float(sheet_lng)
                        elif loc_cell.hyperlink and loc_cell.hyperlink.target:
                            url = loc_cell.hyperlink.target
                            if "google" in url or "goo.gl" in url:
                                total_links += 1
                                latitude, longitude, from_cache = extract_coordinates(url, session, cache)
                                if from_cache:
                                    cached_count += 1
                                elif latitude is not None:
                                    fetched_count += 1
                                    print(f"  ⚡ Fetched new: [{cluster_name}] '{pandal_name}' -> ({latitude}, {longitude})")
                                else:
                                    failed_count += 1
                                    print(f"  ❌ Failed to extract: [{cluster_name}] '{pandal_name}' ({url})")

                        metro_val = sheet.cell(data_r, c_metro).value if c_metro <= sheet.max_column else None
                        station_val = sheet.cell(data_r, c_station).value if c_station <= sheet.max_column else None
                        ferry_val = sheet.cell(data_r, c_ferry).value if c_ferry <= sheet.max_column else None

                        pandal_obj: dict[str, Any] = {
                            "name": pandal_name,
                            "region": sheet_name,
                            "cluster": cluster_name,
                            "location": {
                                "latitude": latitude,
                                "longitude": longitude
                            }
                        }

                        metro_obj = parse_metro(metro_val)
                        if metro_obj:
                            pandal_obj["nearest_metro"] = metro_obj

                        station_obj = parse_station(station_val)
                        if station_obj:
                            pandal_obj["nearest_station"] = station_obj

                        ferry_obj = parse_station(ferry_val)
                        if ferry_obj:
                            pandal_obj["nearest_ferry"] = ferry_obj

                        all_pandals.append(pandal_obj)
                        data_r += 1

    # Save updated cache
    save_cache(cache)

    # Save JSON output
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(all_pandals, f, indent=2, ensure_ascii=False)

    print("\n" + "=" * 40)
    print("           FINAL SUMMARY")
    print("=" * 40)
    print(f"Total Pandals Processed: {len(all_pandals)}")
    print(f"Total Maps Links Found : {total_links}")
    print(f"Loaded From Cache      : {cached_count}")
    print(f"Newly Fetched via Web  : {fetched_count}")
    print(f"Failed to Resolve      : {failed_count}")
    print("=" * 40)
    print(f"\nJSON updated: {OUTPUT_FILE}")
    print(f"Cache saved : {CACHE_FILE}")


if __name__ == "__main__":
    main()